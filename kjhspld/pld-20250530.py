import pandas as pd
import numpy as np
import openpyxl
import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, numbers
from openpyxl.utils import get_column_letter
import logging
import os
from pathlib import Path
import sys
from datetime import datetime

# 配置日志记录 - 同时输出到文件和控制台
def setup_logger():
    """配置日志记录器，同时输出到文件和控制台"""
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # 创建文件处理器
    log_file = Path("accounting_analysis.log")
    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setLevel(logging.INFO)

    # 创建控制台处理器
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)

    # 创建格式化器并添加到处理器
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    # 清除默认处理器并添加自定义处理器
    if logger.hasHandlers():
        logger.handlers.clear()
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger

logger = setup_logger()

class AccountingAnalyzer:
    """会计核算数据与预算执行数据对比分析工具"""

    def __init__(self, yszx_path='./导出数据', kjhs_path='./会计核算.xls', output_path='./财政资金会计核算偏离度.xlsx'):
        """
        初始化分析器

        Args:
            yszx_path: 预算执行数据文件路径
            kjhs_path: 会计核算数据文件路径
            output_path: 输出结果文件路径
        """
        self.yszx_path = Path(yszx_path)
        self.kjhs_path = Path(kjhs_path)
        self.output_path = Path(output_path)
        self.yszx_df = None
        self.kjhs_df = None
        self.merged_df = None
        self.pld_df = None

    def read_data(self):
        """
        读取预算执行数据和会计核算数据

        Returns:
            预算执行数据和会计核算数据的DataFrame元组
        """
        try:
            logger.info("开始读取数据...")

            # 读取预算执行数据
            if (self.yszx_path.with_suffix('.xlsx')).exists():
                yszx_df = pd.read_excel(self.yszx_path.with_suffix('.xlsx'))
                logger.info(f"成功读取预算执行数据: {self.yszx_path.with_suffix('.xlsx')}")
            elif (self.yszx_path.with_suffix('.csv')).exists():
                yszx_df = pd.read_csv(self.yszx_path.with_suffix('.csv'))
                logger.info(f"成功读取预算执行数据: {self.yszx_path.with_suffix('.csv')}")
            else:
                error_msg = f"未找到预算执行数据文件，请检查路径: {self.yszx_path}"
                logger.error(error_msg)
                raise FileNotFoundError(error_msg)

            # 读取会计核算数据
            if self.kjhs_path.exists():
                kjhs_df = pd.read_excel(self.kjhs_path, skiprows=4)
                logger.info(f"成功读取会计核算数据: {self.kjhs_path}")
            else:
                error_msg = f"未找到会计核算数据文件，请检查路径: {self.kjhs_path}"
                logger.error(error_msg)
                raise FileNotFoundError(error_msg)

            self.yszx_df = yszx_df
            self.kjhs_df = kjhs_df
            return yszx_df, kjhs_df

        except Exception as e:
            logger.error(f"读取数据时发生错误: {str(e)}")
            raise

    def process_yszx_data(self):
        """
        处理预算执行数据

        Returns:
            处理后的预算执行数据DataFrame
        """
        if self.yszx_df is None:
            raise ValueError("请先调用read_data方法读取预算执行数据")

        logger.info("开始处理预算执行数据...")

        # 定义指标类型列表
        # 去掉[14]暂存款
        target_types = ['[21]当年预算', '[22]上年结转（非权责制）', '[23]上年结余（非权责制）']

        # 提取政府预算资金，截取编码第一位
        self.yszx_df['资金性质编码'] = pd.to_numeric(self.yszx_df['资金性质'].str.slice(1, 2), errors='coerce')
        # print(self.yszx_df['资金性质编码'])

        # 提取预算单位编码，截取编码第一位
        self.yszx_df['预算单位编码'] = pd.to_numeric(self.yszx_df['预算单位'].str.slice(1, 2), errors='coerce').astype('Int64')
        # print(self.yszx_df['预算单位编码'])

        # 筛选数据
        original_count = len(self.yszx_df)
        self.yszx_df = self.yszx_df[
            (self.yszx_df['预算单位'] != '0') &
            (self.yszx_df['预算单位'] != 0) &
            (self.yszx_df['指标类型'].isin(target_types)) &
            (self.yszx_df['资金性质编码'] == 1) &
            (self.yszx_df['预算单位编码'] != 9)
        ]
        filtered_count = len(self.yszx_df)
        logger.info(f"预算执行数据筛选: 从 {original_count} 行筛选到 {filtered_count} 行")

        # 计算合计金额
        self.yszx_df['预算执行_支出数'] = (
            self.yszx_df['集中支付_实际支出数(非政采)'] +
            self.yszx_df['集中支付_实际支出数（政采）'] +
            self.yszx_df['集中支付_转列支出(非政采)'] +
            self.yszx_df['集中支付_转列支出（政采）'] +
            self.yszx_df['实拨_实际支出']
        )

        # 按预算单位分组并求和
        self.yszx_df = self.yszx_df.groupby('预算单位')['预算执行_支出数'].sum().reset_index()

        # 提取单位编码
        self.yszx_df['单位编码'] = pd.to_numeric(self.yszx_df['预算单位'].str.slice(1, 7), errors='coerce')

        # 添加序号列
        self.yszx_df['序号'] = range(1, len(self.yszx_df) + 1)

        logger.info(f"预算执行数据处理完成，共 {len(self.yszx_df)} 条记录")
        return self.yszx_df

    def process_kjhs_data(self):
        """
        处理会计核算数据

        Returns:
            处理后的会计核算数据DataFrame
        """
        if self.kjhs_df is None:
            raise ValueError("请先调用read_data方法读取会计核算数据")

        logger.info("开始处理会计核算数据...")

        # 过滤掉账套列为借方合计和贷方合计的行
        original_count = len(self.kjhs_df)
        self.kjhs_df = self.kjhs_df[
            (self.kjhs_df['账套'] != '借方合计') &
            (self.kjhs_df['账套'] != '贷方合计')
        ]
        filtered_count = len(self.kjhs_df)
        logger.info(f"会计核算数据筛选: 从 {original_count} 行筛选到 {filtered_count} 行")

        # 去除借方累计列和贷方累计列中字符串里的逗号
        self.kjhs_df['借方累计'] = self.kjhs_df['借方累计'].astype(str).str.replace(',', '')
        self.kjhs_df['贷方累计'] = self.kjhs_df['贷方累计'].astype(str).str.replace(',', '')

        # 将借方累计列和贷方累计列转换为浮点数类型
        self.kjhs_df['借方累计'] = pd.to_numeric(self.kjhs_df['借方累计'], errors='coerce')
        self.kjhs_df['贷方累计'] = pd.to_numeric(self.kjhs_df['贷方累计'], errors='coerce')

        # 将借方累计列和贷方累计列中的 NaN 值转换为 0
        self.kjhs_df['借方累计'] = self.kjhs_df['借方累计'].fillna(0)
        self.kjhs_df['贷方累计'] = self.kjhs_df['贷方累计'].fillna(0)

        # 计算会计核算支出数
        self.kjhs_df['会计核算_支出数'] = (self.kjhs_df['借方累计'] - self.kjhs_df['贷方累计']) / 10000

        # 提取单位编码
        self.kjhs_df['单位编码'] = pd.to_numeric(self.kjhs_df['账套'].str.slice(0, 6), errors='coerce')

        # 按预算单位分组并求和
        self.kjhs_df = self.kjhs_df.groupby('单位编码')['会计核算_支出数'].sum().reset_index()

        logger.info(f"会计核算数据处理完成，共 {len(self.kjhs_df)} 条记录")
        return self.kjhs_df

    def merge_and_analyze(self):
        """
        合并数据并进行差异分析

        Returns:
            包含差异分析结果的DataFrame
        """
        if self.yszx_df is None or self.kjhs_df is None:
            raise ValueError("请先处理预算执行数据和会计核算数据")

        logger.info("开始合并数据并进行差异分析...")

        # 合并两个 DataFrame
        self.merged_df = pd.merge(self.yszx_df, self.kjhs_df, on='单位编码', how='left')

        # 将会计核算_支出数列中的空值填充为 0
        self.merged_df['会计核算_支出数'] = self.merged_df['会计核算_支出数'].fillna(0)

        # 计算差额和偏离度
        self.merged_df['差额'] = self.merged_df['会计核算_支出数'] - self.merged_df['预算执行_支出数']

        # 安全计算偏离度，处理除零错误
        def calculate_deviation(row):
            if pd.notna(row['预算执行_支出数']) and row['预算执行_支出数'] != 0:
                return row['差额'] / row['预算执行_支出数']
            return np.nan

        self.merged_df['偏离度'] = self.merged_df.apply(calculate_deviation, axis=1)

        # 处理差额的精度问题，四舍五入保留 6 位小数
        self.merged_df['差额'] = self.merged_df['差额'].round(6)

        self.merged_df['备注'] =''
        # 调整列顺序
        self.pld_df = self.merged_df[['序号', '单位编码', '预算单位', '预算执行_支出数', '会计核算_支出数', '差额', '偏离度', '备注']]

        logger.info(f"数据合并和分析完成，共 {len(self.pld_df)} 条记录")
        return self.pld_df

    def save_and_format_excel(self):
        """
        将结果保存到Excel并进行格式美化
        """
        if self.pld_df is None:
            raise ValueError("请先进行数据合并和分析")

        try:
            logger.info(f"开始保存结果到Excel: {self.output_path}")

            # 确保输出目录存在
            self.output_path.parent.mkdir(parents=True, exist_ok=True)

            # 将结果保存到 Excel 文件
            self.pld_df.to_excel(self.output_path, sheet_name='偏离度', index=False)

            # 加载保存的 Excel 文件进行格式美化
            wb = load_workbook(self.output_path)
            ws = wb['偏离度']

            # 定义样式
            # 边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 表头样式
            header_font = Font(name='微软雅黑', bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

            # 高偏离度行样式
            high_deviation_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            # 为所有单元格添加边框
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border

            # 动态调整列宽
            column_widths = []
            for row in ws.iter_rows():
                for i, cell in enumerate(row):
                    try:
                        if len(column_widths) > i:
                            if len(str(cell.value)) > column_widths[i]:
                                column_widths[i] = len(str(cell.value))
                        else:
                            column_widths += [len(str(cell.value))]
                    except:
                        pass

            for i, column_width in enumerate(column_widths):
                ws.column_dimensions[get_column_letter(i + 1)].width = min(column_width + 2, 30)

            # 美化表头
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                cell.font = header_font
                cell.fill = header_fill

            # 找到偏离度列的索引
            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            deviation_index = header.index('偏离度')

            # 格式化数据行
            for row in ws.iter_rows(min_row=2):
                deviation_cell = row[deviation_index]
                try:
                    # 直接获取单元格的数值
                    deviation = deviation_cell.value
                    if pd.notna(deviation):
                        # 设置单元格格式为百分比，保留两位小数
                        deviation_cell.number_format = '0.00%'
                        # 高偏离度行标黄
                        if abs(deviation) > 0.1:
                            for cell in row:
                                cell.fill = high_deviation_fill
                except (ValueError, AttributeError):
                    continue

            # 第一行固定（冻结窗格）
            ws.freeze_panes = ws['A2']
            
            # 设置筛选器
            max_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f'A1:{max_col}{ws.max_row}'

            # 添加汇总信息
            total_row = len(self.pld_df) + 2
            ws[f'A{total_row}'] = '汇总统计'
            ws[f'A{total_row}'].font = Font(bold=True)

            ws[f'B{total_row}'] = '总预算单位数'
            ws[f'C{total_row}'] = len(self.pld_df)

            ws[f'B{total_row + 1}'] = '高于偏离度10%的预算单位数'
            ws[f'C{total_row + 1}'] = f'=COUNTIF({get_column_letter(deviation_index + 1)}2:{get_column_letter(deviation_index + 1)}{total_row - 1},">0.1")+COUNTIF({get_column_letter(deviation_index + 1)}2:{get_column_letter(deviation_index + 1)}{total_row - 1},"<-0.1")'

            # 为汇总行添加边框
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                for row in range(total_row, total_row + 2):
                    ws[f'{col}{row}'].border = thin_border

            # 添加生成时间戳
            ws[f'A{total_row + 2}'] = f'时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'

            # 保存修改后的 Excel 文件
            wb.save(self.output_path)
            logger.info(f"Excel文件格式化完成: {self.output_path}")
            print(f"分析结果已成功保存到: {self.output_path}")

        except Exception as e:
            logger.error(f"保存和格式化Excel文件时发生错误: {str(e)}")
            raise

    def run_analysis(self):
        """执行完整的分析流程"""
        try:
            logger.info("=" * 50)
            logger.info("开始会计核算与预算执行数据对比分析")
            logger.info("=" * 50)

            self.read_data()
            self.process_yszx_data()
            self.process_kjhs_data()
            self.merge_and_analyze()
            self.save_and_format_excel()

            logger.info("=" * 50)
            logger.info("分析完成")
            logger.info("=" * 50)

        except Exception as e:
            logger.error(f"分析过程中发生致命错误: {str(e)}")
            print(f"程序执行失败: {str(e)}")
            sys.exit(1)


if __name__ == "__main__":
    # 创建分析器实例并执行分析
    analyzer = AccountingAnalyzer()
    analyzer.run_analysis()    