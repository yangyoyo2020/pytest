import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional, Tuple, Dict, Any
import warnings
from contextlib import contextmanager

# 抑制pandas警告
warnings.filterwarnings('ignore', category=pd.errors.PerformanceWarning)

class ConfigManager:
    """配置管理类"""
    
    # 默认配置
    DEFAULT_CONFIG = {
        'target_types': ['[21]当年预算', '[22]上年结转（非权责制）', '[23]上年结余（非权责制）'],
        'fund_nature_code': 1,  # 政府预算资金编码
        'excluded_unit_code': 9,  # 排除的预算单位编码
        'high_deviation_threshold': 0.1,  # 高偏离度阈值
        'decimal_places': 6,  # 差额保留小数位数
        'percentage_format': '0.00%',  # 百分比格式
        'max_column_width': 30,  # 最大列宽
        'column_width_padding': 2,  # 列宽填充
    }
    
    @classmethod
    def get_config(cls, key: str, default=None):
        """获取配置值"""
        return cls.DEFAULT_CONFIG.get(key, default)

class Logger:
    """日志管理类"""
    
    _instance = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._setup_logger()
        return cls._instance
    
    def _setup_logger(self):
        """配置日志记录器"""
        self.logger = logging.getLogger('AccountingAnalyzer')
        self.logger.setLevel(logging.INFO)
        
        # 避免重复添加处理器
        if self.logger.hasHandlers():
            self.logger.handlers.clear()
        
        # 文件处理器
        log_file = Path("accounting_analysis.log")
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        
        # 控制台处理器
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        
        # 格式化器
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)
        
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
    
    def info(self, message: str):
        self.logger.info(message)
    
    def error(self, message: str):
        self.logger.error(message)
    
    def warning(self, message: str):
        self.logger.warning(message)

class DataProcessor:
    """数据处理基类"""
    
    def __init__(self, logger: Logger):
        self.logger = logger
    
    @staticmethod
    def safe_numeric_conversion(series: pd.Series, errors: str = 'coerce') -> pd.Series:
        """安全的数值转换"""
        return pd.to_numeric(series, errors=errors)
    
    @staticmethod
    def clean_numeric_string(series: pd.Series) -> pd.Series:
        """清理数值字符串（去除逗号等）"""
        return series.astype(str).str.replace(',', '').str.replace(' ', '')
    
    def validate_dataframe(self, df: pd.DataFrame, required_columns: list) -> bool:
        """验证DataFrame是否包含必需的列"""
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            self.logger.error(f"缺少必需的列: {missing_columns}")
            return False
        return True

class BudgetExecutionProcessor(DataProcessor):
    """预算执行数据处理器"""
    
    REQUIRED_COLUMNS = ['预算单位', '指标类型', '资金性质', '集中支付_实际支出数(非政采)', 
                       '集中支付_实际支出数（政采）', '集中支付_转列支出(非政采)', 
                       '集中支付_转列支出（政采）', '实拨_实际支出']
    
    def process(self, df: pd.DataFrame) -> pd.DataFrame:
        """处理预算执行数据"""
        if not self.validate_dataframe(df, self.REQUIRED_COLUMNS):
            raise ValueError("预算执行数据格式不正确")
        
        self.logger.info("开始处理预算执行数据...")
        original_count = len(df)
        
        # 数据清洗和转换
        df = self._clean_and_transform(df)
        
        # 数据筛选
        df = self._filter_data(df)
        
        # 计算支出数
        df = self._calculate_expenditure(df)
        
        # 分组汇总
        df = self._group_and_aggregate(df)
        
        # 添加辅助列
        df = self._add_helper_columns(df)
        
        filtered_count = len(df)
        self.logger.info(f"预算执行数据处理完成: {original_count} -> {filtered_count} 条记录")
        
        return df
    
    def _clean_and_transform(self, df: pd.DataFrame) -> pd.DataFrame:
        """清洗和转换数据"""
        # 提取资金性质编码
        df['资金性质编码'] = self.safe_numeric_conversion(
            df['资金性质'].str.slice(1, 2)
        )
        
        # 提取预算单位编码
        df['预算单位编码'] = self.safe_numeric_conversion(
            df['预算单位'].str.slice(1, 2)
        ).astype('Int64')
        
        return df
    
    def _filter_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """筛选数据"""
        target_types = ConfigManager.get_config('target_types')
        fund_nature_code = ConfigManager.get_config('fund_nature_code')
        excluded_unit_code = ConfigManager.get_config('excluded_unit_code')
        
        return df[
            (df['预算单位'] != '0') &
            (df['预算单位'] != 0) &
            (df['指标类型'].isin(target_types)) &
            (df['资金性质编码'] == fund_nature_code) &
            (df['预算单位编码'] != excluded_unit_code)
        ].copy()
    
    def _calculate_expenditure(self, df: pd.DataFrame) -> pd.DataFrame:
        """计算支出数"""
        expenditure_columns = [
            '集中支付_实际支出数(非政采)', '集中支付_实际支出数（政采）',
            '集中支付_转列支出(非政采)', '集中支付_转列支出（政采）', '实拨_实际支出'
        ]
        
        # 确保所有列都是数值类型
        for col in expenditure_columns:
            df[col] = self.safe_numeric_conversion(df[col]).fillna(0)
        
        df['预算执行_支出数'] = df[expenditure_columns].sum(axis=1)
        return df
    
    def _group_and_aggregate(self, df: pd.DataFrame) -> pd.DataFrame:
        """分组并汇总"""
        return df.groupby('预算单位', as_index=False)['预算执行_支出数'].sum()
    
    def _add_helper_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """添加辅助列"""
        df['单位编码'] = self.safe_numeric_conversion(
            df['预算单位'].str.slice(1, 7)
        )
        df['序号'] = range(1, len(df) + 1)
        return df

class AccountingProcessor(DataProcessor):
    """会计核算数据处理器"""
    
    REQUIRED_COLUMNS = ['账套', '借方累计', '贷方累计']
    
    def process(self, df: pd.DataFrame) -> pd.DataFrame:
        """处理会计核算数据"""
        if not self.validate_dataframe(df, self.REQUIRED_COLUMNS):
            raise ValueError("会计核算数据格式不正确")
        
        self.logger.info("开始处理会计核算数据...")
        original_count = len(df)
        
        # 数据清洗
        df = self._clean_data(df)
        
        # 数值转换
        df = self._convert_numeric_columns(df)
        
        # 计算支出数
        df = self._calculate_expenditure(df)
        
        # 提取单位编码并分组
        df = self._extract_unit_code_and_group(df)
        
        filtered_count = len(df)
        self.logger.info(f"会计核算数据处理完成: {original_count} -> {filtered_count} 条记录")
        
        return df
    
    def _clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """清洗数据"""
        return df[
            (df['账套'] != '借方合计') & 
            (df['账套'] != '贷方合计')
        ].copy()
    
    def _convert_numeric_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """转换数值列"""
        for col in ['借方累计', '贷方累计']:
            # 清理字符串并转换为数值
            df[col] = self.safe_numeric_conversion(
                self.clean_numeric_string(df[col])
            ).fillna(0)
        
        return df
    
    def _calculate_expenditure(self, df: pd.DataFrame) -> pd.DataFrame:
        """计算支出数"""
        df['会计核算_支出数'] = (df['借方累计'] - df['贷方累计']) / 10000
        return df
    
    def _extract_unit_code_and_group(self, df: pd.DataFrame) -> pd.DataFrame:
        """提取单位编码并分组"""
        df['单位编码'] = self.safe_numeric_conversion(
            df['账套'].str.slice(0, 6)
        )
        
        return df.groupby('单位编码', as_index=False)['会计核算_支出数'].sum()

class ExcelFormatter:
    """Excel格式化工具"""
    
    def __init__(self, logger: Logger):
        self.logger = logger
    
    def format_excel(self, file_path: Path, sheet_name: str, data_rows: int):
        """格式化Excel文件"""
        try:
            wb = load_workbook(file_path)
            ws = wb[sheet_name]
            
            self._apply_styles(ws)
            self._adjust_column_widths(ws)
            self._format_headers(ws)
            self._format_data_rows(ws, data_rows)
            self._add_features(ws, data_rows)
            self._add_summary_statistics(ws, data_rows)
            
            wb.save(file_path)
            self.logger.info(f"Excel格式化完成: {file_path}")
            
        except Exception as e:
            self.logger.error(f"Excel格式化失败: {str(e)}")
            raise
    
    def _apply_styles(self, ws):
        """应用样式"""
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 为所有单元格添加边框
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    def _adjust_column_widths(self, ws):
        """调整列宽"""
        max_column_width = ConfigManager.get_config('max_column_width')
        column_width_padding = ConfigManager.get_config('column_width_padding')
        
        column_widths = {}
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                width = len(str(cell.value or ''))
                column_widths[i] = max(column_widths.get(i, 0), width)
        
        for i, width in column_widths.items():
            column_letter = get_column_letter(i + 1)
            ws.column_dimensions[column_letter].width = min(
                width + column_width_padding, max_column_width
            )
    
    def _format_headers(self, ws):
        """格式化表头"""
        header_font = Font(name='微软雅黑', bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            cell.font = header_font
            cell.fill = header_fill
    
    def _format_data_rows(self, ws, data_rows: int):
        """格式化数据行"""
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        
        try:
            deviation_index = header.index('偏离度')
        except ValueError:
            self.logger.warning("未找到偏离度列，跳过偏离度格式化")
            return
        
        high_deviation_fill = PatternFill(
            start_color='FFFF00', end_color='FFFF00', fill_type='solid'
        )
        percentage_format = ConfigManager.get_config('percentage_format')
        high_deviation_threshold = ConfigManager.get_config('high_deviation_threshold')
        
        for row in ws.iter_rows(min_row=2, max_row=data_rows + 1):
            deviation_cell = row[deviation_index]
            
            if deviation_cell.value is not None:
                deviation_cell.number_format = percentage_format
                
                # 高偏离度行标黄
                if abs(deviation_cell.value) > high_deviation_threshold:
                    for cell in row:
                        cell.fill = high_deviation_fill
    
    def _add_features(self, ws, data_rows: int):
        """添加Excel功能"""
        # 冻结窗格
        ws.freeze_panes = ws['A2']
        
        # 自动筛选
        max_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f'A1:{max_col}{data_rows + 1}'
    
    def _add_summary_statistics(self, ws, data_rows: int):
        """添加汇总统计"""
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        
        try:
            deviation_index = header.index('偏离度')
            deviation_col = get_column_letter(deviation_index + 1)
        except ValueError:
            self.logger.warning("未找到偏离度列，跳过统计信息")
            return
        
        summary_row = data_rows + 3
        high_deviation_threshold = ConfigManager.get_config('high_deviation_threshold')
        
        # 添加统计信息
        statistics = [
            ('汇总统计', '', ''),
            ('总预算单位数', '', data_rows),
            ('高于偏离度10%的预算单位数', '', 
             f'=COUNTIF({deviation_col}2:{deviation_col}{data_rows + 1},">{high_deviation_threshold}")'
             f'+COUNTIF({deviation_col}2:{deviation_col}{data_rows + 1},"<{-high_deviation_threshold}")'),
            (f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', '', '')
        ]
        
        for i, (col_a, col_b, col_c) in enumerate(statistics):
            row_num = summary_row + i
            ws[f'A{row_num}'] = col_a
            ws[f'B{row_num}'] = col_b
            ws[f'C{row_num}'] = col_c
            
            if i == 0:  # 标题行加粗
                ws[f'A{row_num}'].font = Font(bold=True)

class AccountingAnalyzer:
    """会计核算数据与预算执行数据对比分析工具"""
    
    def __init__(self, yszx_path: str = './导出数据', 
                 kjhs_path: str = './会计核算.xls', 
                 output_path: str = './财政资金会计核算偏离度.xlsx'):
        """初始化分析器"""
        self.yszx_path = Path(yszx_path)
        self.kjhs_path = Path(kjhs_path)
        self.output_path = Path(output_path)
        
        self.logger = Logger()
        self.budget_processor = BudgetExecutionProcessor(self.logger)
        self.accounting_processor = AccountingProcessor(self.logger)
        self.excel_formatter = ExcelFormatter(self.logger)
        
        self.yszx_df = None
        self.kjhs_df = None
        self.result_df = None
    
    @contextmanager
    def error_handler(self, operation: str):
        """错误处理上下文管理器"""
        try:
            yield
        except Exception as e:
            self.logger.error(f"{operation}时发生错误: {str(e)}")
            raise
    
    def read_data(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """读取数据文件"""
        with self.error_handler("读取数据"):
            self.logger.info("开始读取数据...")
            
            # 读取预算执行数据
            yszx_df = self._read_budget_data()
            
            # 读取会计核算数据
            kjhs_df = self._read_accounting_data()
            
            self.yszx_df = yszx_df
            self.kjhs_df = kjhs_df
            
            return yszx_df, kjhs_df
    
    def _read_budget_data(self) -> pd.DataFrame:
        """读取预算执行数据"""
        for suffix in ['.xlsx', '.csv']:
            file_path = self.yszx_path.with_suffix(suffix)
            if file_path.exists():
                if suffix == '.xlsx':
                    df = pd.read_excel(file_path)
                else:
                    df = pd.read_csv(file_path)
                self.logger.info(f"成功读取预算执行数据: {file_path}")
                return df
        
        raise FileNotFoundError(f"未找到预算执行数据文件: {self.yszx_path}")
    
    def _read_accounting_data(self) -> pd.DataFrame:
        """读取会计核算数据"""
        if not self.kjhs_path.exists():
            raise FileNotFoundError(f"未找到会计核算数据文件: {self.kjhs_path}")
        
        df = pd.read_excel(self.kjhs_path, skiprows=4)
        self.logger.info(f"成功读取会计核算数据: {self.kjhs_path}")
        return df
    
    def process_data(self) -> pd.DataFrame:
        """处理数据"""
        with self.error_handler("处理数据"):
            if self.yszx_df is None or self.kjhs_df is None:
                raise ValueError("请先读取数据")
            
            # 处理预算执行数据
            processed_yszx = self.budget_processor.process(self.yszx_df)
            
            # 处理会计核算数据
            processed_kjhs = self.accounting_processor.process(self.kjhs_df)
            
            # 合并数据并计算偏离度
            result_df = self._merge_and_calculate_deviation(processed_yszx, processed_kjhs)
            
            self.result_df = result_df
            return result_df
    
    def _merge_and_calculate_deviation(self, yszx_df: pd.DataFrame, 
                                     kjhs_df: pd.DataFrame) -> pd.DataFrame:
        """合并数据并计算偏离度"""
        self.logger.info("开始合并数据并计算偏离度...")
        
        # 合并数据
        merged_df = pd.merge(yszx_df, kjhs_df, on='单位编码', how='left')
        merged_df['会计核算_支出数'] = merged_df['会计核算_支出数'].fillna(0)
        
        # 计算差额和偏离度
        merged_df['差额'] = merged_df['会计核算_支出数'] - merged_df['预算执行_支出数']
        merged_df['差额'] = merged_df['差额'].round(
            ConfigManager.get_config('decimal_places')
        )
        
        # 安全计算偏离度
        merged_df['偏离度'] = merged_df.apply(self._calculate_deviation_safe, axis=1)
        merged_df['备注'] = ''
        
        # 重新排列列
        result_df = merged_df[[
            '序号', '单位编码', '预算单位', '预算执行_支出数', 
            '会计核算_支出数', '差额', '偏离度', '备注'
        ]]
        
        self.logger.info(f"数据合并完成，共 {len(result_df)} 条记录")
        return result_df
    
    @staticmethod
    def _calculate_deviation_safe(row) -> Optional[float]:
        """安全计算偏离度"""
        if pd.notna(row['预算执行_支出数']) and row['预算执行_支出数'] != 0:
            return row['差额'] / row['预算执行_支出数']
        return None
    
    def save_results(self):
        """保存结果"""
        with self.error_handler("保存结果"):
            if self.result_df is None:
                raise ValueError("请先处理数据")
            
            self.logger.info(f"开始保存结果到: {self.output_path}")
            
            # 确保输出目录存在
            self.output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # 保存到Excel
            self.result_df.to_excel(self.output_path, sheet_name='偏离度', index=False)
            
            # 格式化Excel
            self.excel_formatter.format_excel(
                self.output_path, '偏离度', len(self.result_df)
            )
            
            print(f"分析结果已保存到: {self.output_path}")
    
    def run_analysis(self):
        """执行完整分析流程"""
        try:
            self.logger.info("=" * 50)
            self.logger.info("开始会计核算与预算执行数据对比分析")
            self.logger.info("=" * 50)
            
            self.read_data()
            self.process_data()
            self.save_results()
            
            self.logger.info("=" * 50)
            self.logger.info("分析完成")
            self.logger.info("=" * 50)
            
        except Exception as e:
            self.logger.error(f"分析失败: {str(e)}")
            raise

def main():
    """主函数"""
    try:
        analyzer = AccountingAnalyzer()
        analyzer.run_analysis()
    except Exception as e:
        print(f"程序执行失败: {str(e)}")
        return 1
    return 0

if __name__ == "__main__":
    import sys
    sys.exit(main())